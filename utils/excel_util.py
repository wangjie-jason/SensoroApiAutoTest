#!/usr/bin/python
# -*- coding:utf-8 -*-
# @Time : 2023/7/6 20:49
# @Author : wangjie
# @File : excel_util.py
# @project : SensoroApiAutoTest

from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.logger import logger


class ExcelUtil:
    """excel文件相关操作"""

    @staticmethod
    def create_excel(filename: str) -> str:
        """创建excel文件，需要指定excel文件的绝对路径，如D:\test\test.xlsx"""
        wb = openpyxl.Workbook()
        wb.save(filename)
        wb.close()
        logger.info(f"创建excel文件: {filename}")
        return filename

    @staticmethod
    def read_sheet(sheet_name: str, workbook: Workbook) -> dict[str, Any]:
        """读取指定表单的内容"""
        ws: Worksheet = workbook[sheet_name]
        all_values = list(ws.values)
        header = all_values[0]
        return {
            "sheet_name": sheet_name,
            "data": [dict(zip(header, row)) for row in all_values[1:]]
        }

    @staticmethod
    def read(filename: str, sheet: str | None = None) -> list[dict[str, Any]]:
        """读取excel数据并返回，sheet为空则读取所有表单"""
        workbook = openpyxl.load_workbook(filename)
        sheets = workbook.sheetnames
        results: list[dict[str, Any]] = []

        if sheet:
            if sheet not in sheets:
                workbook.close()
                raise ValueError(f"表单【{sheet}】不存在，可用表单: {sheets}")
            results.append(ExcelUtil.read_sheet(sheet, workbook))
        else:
            for s in sheets:
                results.append(ExcelUtil.read_sheet(s, workbook))

        workbook.close()
        return results

    @staticmethod
    def write(filename: str, row: int, column: int, data: Any, sheet_name: str | None = None) -> None:
        """往excel写入数据，sheet_name为空则使用活动表单"""
        workbook = openpyxl.load_workbook(filename)
        sheets = workbook.sheetnames

        if sheet_name:
            if sheet_name not in sheets:
                workbook.close()
                raise ValueError(f"表单【{sheet_name}】不存在，可用表单: {sheets}")
            ws: Worksheet = workbook[sheet_name]
            logger.info(f"往表单【{sheet_name}】中写入数据")
        else:
            ws = workbook.active
            logger.info(f"未指定表单，默认使用活动表单【{ws.title}】")

        ws.cell(row=row, column=column, value=data)
        workbook.save(filename)
        workbook.close()